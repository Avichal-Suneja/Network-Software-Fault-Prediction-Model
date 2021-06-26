import openpyxl;

allWorkbook = openpyxl.load_workbook("allPacketData.xlsx")
errorWorkbook = openpyxl.load_workbook("errorPacketData.xlsx")

newWorkbook = openpyxl.load_workbook("datasheet.xlsx")
newSheet = newWorkbook[newWorkbook.sheetnames[0]]

allSheet = allWorkbook[allWorkbook.sheetnames[0]]
errorSheet = errorWorkbook[errorWorkbook.sheetnames[0]]

errorIndexes = []

NUM = 1
TIME = 2
PROTOCOL = 5
LENGTH = 6

for i in range(2,errorSheet.max_row):
    cell = errorSheet.cell(row=i, column=1)
    errorIndexes.append(cell.value)

for i in range(2,allSheet.max_row):
    num = allSheet.cell(row=i, column=NUM)
    time = allSheet.cell(row=i, column=TIME)
    protocol = allSheet.cell(row=i, column=PROTOCOL)
    len = allSheet.cell(row=i, column=LENGTH)

    newSheet.cell(row=i, column=1).value = time.value
    newSheet.cell(row=i, column=2).value = protocol.value
    newSheet.cell(row=i, column=3).value = len.value
    newSheet.cell(row=i, column=4).value = int((num.value in errorIndexes) == True)

for i in range(3, newSheet.max_row):
    time = newSheet.cell(row=i, column=1)
    prevTime = newSheet.cell(row=i-1, column=1)
    delTime = newSheet.cell(row=i, column=5)

    delTime.value = time.value - prevTime.value


newWorkbook.save("datasheet.xlsx")




