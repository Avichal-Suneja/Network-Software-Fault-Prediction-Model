import openpyxl

workbook = openpyxl.load_workbook("datasheet.xlsx")
sheet = workbook[workbook.sheetnames[0]]

protocolList = {
    'TCP' : 0,
    'SSDP' : 1,
    'DNS' : 2,
    'HTTP' : 3,
    'MDNS' : 4,
    'LLMNR' : 5,
    'NBNS' : 6,
    'TLSv1.3' : 7,
    'TLSv1.2' : 8,
    'QUIC' : 9,
    'IGMPv3' : 10,
    'ICMPv6' : 11,
    'ARP' : 12,
    'ICMP' : 13,
}

for i in range(2, sheet.max_row):
    protocol = sheet.cell(row=i, column=2)
    protocol.value = protocolList[protocol.value]

workbook.save("datasheet.xlsx")